from tkinter import Tk, Button, Label, filedialog, messagebox
import openpyxl
import pandas as pd
import time
# Tkinter
window = Tk()

# Cấu hình
window.title("Remake Project")
window.geometry("300x200")

# Label
result_label = Label(window, text="")
result_label.pack()

def process_data(file_path):
    start_time = time.time()
    workbook = openpyxl.load_workbook(file_path)
    sheetTongHop = workbook['TongHop']

    data = pd.read_excel(file_path, sheet_name=None)

    DanhMuc = data['DanhMuc']
    DonGia = data['DonGia']
    BanHang = data['BanHang']
    TongHop = data['TongHop']

    merged_data = BanHang.merge(DanhMuc, on='Ma_So', how='inner')

    totalPay = []
    quantity_column = getQuantityColumn(BanHang, 'Thoi_Gian')
    don_gia_quantity_column = getQuantityColumn(DonGia, 'Thoi_Gian')

    period_data = dict(zip(TongHop['Ma_So'], zip(TongHop['Tu_Ngay'], TongHop['Den_Ngay'])))

    for index, row in merged_data.iterrows():
        sheetTongHop.cell(row=index + 2, column=2).value = row['Ten_Hang']

        count = 0
        for i in range(1, quantity_column + 1):
            period_start, period_end = getPeriod(row['Ma_So'], period_data)
            if period_start <= row['Thoi_Gian{}'.format(i)] <= period_end:
                if str(row['So_Luong{}'.format(i)]).strip() != '-':
                    count += int(row['So_Luong{}'.format(i)])
        sheetTongHop.cell(row=index + 2, column=5).value = count

        matching_rows = DonGia[DonGia['Ma_So'] == row['Ma_So']]
        if not matching_rows.empty:
            totalPayProduct = 0
            for i in range(1, quantity_column * 2, 2):
                for j in range(don_gia_quantity_column * 2, 1, -2):
                    time_range = getPeriod(row['Ma_So'], period_data)
                    if row[i] >= matching_rows.iloc[0, j] and time_range[0] <= row[i] <= time_range[1]:
                        quantity = str(row[i + 1]).strip()
                        if quantity != '-':
                            totalPayProduct += int(quantity) * int(matching_rows.iloc[0, j - 1])
                            break
            totalPay.append(totalPayProduct)
            sheetTongHop.cell(row=index + 2, column=6).value = "{:,.0f}".format(totalPayProduct)

    sheetTongHop.cell(row=len(merged_data) + 2, column=6).value = "{:,.0f}".format(sum(totalPay))

    workbook.save(file_path)
    workbook.close()

    end_time = time.time()
    elapsed_time = end_time - start_time
    result_label.config(text=f"Thời gian: {elapsed_time:.2f} giây")
    messagebox.showinfo("Success", "Dữ lịu đã được xử lý")

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        result_label.config(text="Đang xử lý...")
        window.update()
        process_data(file_path)
def getPeriod(target, period_data):
    return period_data[target]
def getQuantityColumn(worksheet, target):
    column_names = worksheet.columns.tolist()
    filtered_columns = [col for col in column_names if target in col]
    return len(filtered_columns)
# Nút
select_button = Button(window, text="Chọn tệp cần xử lý", command=select_file)
select_button.pack(pady=20)

# Loop
window.mainloop()
